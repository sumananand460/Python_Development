from io import BytesIO
from collections import defaultdict
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count, Q
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from .models import Attendance, Student, SchoolClass, Notice

def monthly_attendance_stats(student):
    records = student.attendance_records.all()
    present = records.filter(status=Attendance.PRESENT).count()
    absent = records.filter(status=Attendance.ABSENT).count()
    total = present + absent
    percentage = round((present / total) * 100, 2) if total else 0.0
    return {"present": present, "absent": absent, "percentage": percentage, "total": total}

def export_attendance_excel(rows, title="Attendance Report"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"
    ws.append([title])
    ws.append([])
    for row in rows:
        ws.append(list(row))
    for cell in ws[1]:
        cell.font = Font(bold=True, size=14)
    if ws.max_row >= 3:
        for cell in ws[3]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
    for col in ws.columns:
        max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = HttpResponse(stream.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="attendance_report.xlsx"'
    return response

def export_attendance_pdf(rows, title="Attendance Report"):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles["Title"]), Spacer(1, 12)]
    data = [list(rows[0])] + [list(r) for r in rows[1:]] if rows else [[title]]
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightblue),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="attendance_report.pdf"'
    return response
